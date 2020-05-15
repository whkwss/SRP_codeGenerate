from openpyxl import load_workbook

def excel2txt(excelpath, txtpath):
    print('读取Excel...')
    excelpath = excelpath
    txtpath = txtpath

    book = load_workbook(excelpath)
    # statecolumns =stateSheet.max_column
    sheets = []
    columns = []
    for item in book.sheetnames:
        ws = book[item]
        sheets.append(list(ws))
        columns.append(ws.max_column)
        ws.delete_rows(0)

    print('写入Txt...')
    txt = open(txtpath, 'w')

    for i in range(len(sheets)):
        for row in sheets[i]:
            if row == sheets[i][0]:
                continue
            data = ''
            for j in range(columns[i]):
                if j != columns[i]-1:
                    data += str(row[j].value) + '\t'
                else:
                    if i == len(sheets)-1 and row == sheets[i][-1]:
                        data += str(row[j].value)
                    else:
                        data += str(row[j].value) + '\n'
            txt.write(data)
    txt.close()
    print('Txt写入完毕')

if __name__ == '__main__':
    excelpath = 'testData.xlsx'
    txtpath = 'excel2txt.txt'
    excel2txt(excelpath, txtpath)
